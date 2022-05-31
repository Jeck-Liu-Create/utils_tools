# -*- encoding: utf-8 -*-
'''
@File    :   from2.py    
@Contact :  tclyldx@163.com
@License :   (C)Copyright LiuDongxing,All Rights Reserved. 
             Unauthorized copying of this file, via any medium is strictly prohibited 
             Proprietary and confidential
@Modify Time      @Author    @Version    @Desciption
------------      -------    --------    -----------
2022/5/30 20:16   LiuDongxing      1.0         None
'''

# import lib
import openpyxl
wb = openpyxl.load_workbook('test2.xlsx')
name_list = wb.sheetnames
for sheetName in name_list:
    mergedRowCount = 0
    dw_sheet = wb[sheetName]
    startDataRow = 2  # 从第三行开始写入输出报告的数据
    needGroupAndMergeColumnsIndex = [9]  # 输出报告的前三列需要合并字母
    for columnNumber in needGroupAndMergeColumnsIndex:
        # 当前单元格的值
        currentValue = 0
        mergeValue = 0
        firstMergeRowNumber = 2
        # 合并单元格的个数
        addFlag = 0
        for rowNumber in range(startDataRow, dw_sheet.max_row + 2):
            mergeValue = currentValue
            currentValue = dw_sheet.cell(
                row=rowNumber, column=columnNumber).value

            if currentValue == mergeValue:
                addFlag += 1
                if addFlag == 1:
                    firstMergeRowNumber = rowNumber - 1
            else:
                if addFlag == 0:
                    firstMergeRowNumber = rowNumber
                    mergedRowCount += 1
                else:
                    # mergedRowCount += 1
                    # dw_sheet.merge_cells(
                    #     start_row=firstMergeRowNumber,
                    #     end_row=rowNumber - 1,
                    #     start_column=columnNumber,
                    #     end_column=columnNumber)
                    # dw_sheet.cell(row=firstMergeRowNumber,
                    #               column=columnNumber).value = mergeValue
                    # if columnNumber == 1:

                        # tempValueCol5 = 0
                        # #
                        #
                        # if sheetName == name_list[0]:
                        #     tempValueCol5 = dw_sheet.cell(
                        #         row=firstMergeRowNumber, column=5).value
                        # else:
                        #     tempValueCol5 = dw_sheet.cell(
                        #         row=firstMergeRowNumber, column=4).value * 0.8 - dw_sheet.cell(
                        #         row=firstMergeRowNumber, column=4).value
                        #
                        # #
                        # # 合并单元格
                        # dw_sheet.merge_cells(
                        #     start_row=firstMergeRowNumber,
                        #     end_row=rowNumber - 1,
                        #     start_column=5,
                        #     end_column=5)
                        # dw_sheet.cell(row=firstMergeRowNumber,
                        #               column=5).value = tempValueCol5

                    for merge_col in [10,11,12,13,14]:
                        dw_sheet.merge_cells(
                            start_row=firstMergeRowNumber,
                            end_row=rowNumber - 1,
                            start_column=merge_col,
                            end_column=merge_col)
                    #

                    # 画单元格的界线
                    #                             if addFlag == 1:
                    #                                  for col in range(1,5):
                    #                                         dw_sheet.cell(row=firstMergeRowNumber+1,column=col).border=border
                    #                             else:
                    #                                 for row in range(firstMergeRowNumber+1,rowNumber):
                    #                                     for col in range(1,5):
                    #                                         dw_sheet.cell(row=row,column=col).border=border
                    addFlag = 0
                    print(firstMergeRowNumber)
        # if columnNumber == 1:
        #     totalFormattedRowsCount = mergedRowCount - 1
        #     raw_dw_txt = dw_sheet.cell(row=1, column=1).value
        #     raw_dw_cnt = '（' + str(totalFormattedRowsCount)
        #     dw_txt = raw_dw_txt + raw_dw_cnt
        #     dw_sheet['A1'] = dw_txt
            # wb.save('./template.xlsx')
    wb.save('test2.xlsx')
    wb.close()